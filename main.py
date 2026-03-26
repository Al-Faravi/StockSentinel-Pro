"""
╔══════════════════════════════════════════════════════════════════╗
║              StockSentinel Pro                                   ║
║         Comprehensive Inventory Management System                ║
║                                                                  ║
║  Tech Stack: Python · Tkinter · SQLite                           ║
║  Features: Role-Based Access · SQL Relational DB                 ║
║            Search & Filter · Excel/PDF Export                    ║
╚══════════════════════════════════════════════════════════════════╝
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
import hashlib
import os
import datetime

# ── optional deps ──────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    PDF_OK = True
except ImportError:
    PDF_OK = False

DB_FILE = "stocksentinel.db"

# ══════════════════════════════════════════════════════════════════
#  COLOUR PALETTE
# ══════════════════════════════════════════════════════════════════
C = {
    "bg":        "#0D1117",
    "surface":   "#161B22",
    "surface2":  "#21262D",
    "border":    "#30363D",
    "accent":    "#238636",
    "accent_h":  "#2EA043",
    "danger":    "#DA3633",
    "warning":   "#D29922",
    "info":      "#1F6FEB",
    "text":      "#E6EDF3",
    "muted":     "#8B949E",
    "highlight": "#388BFD",
    "white":     "#FFFFFF",
    "tag_bg":    "#1C2128",
}

FONT_HEAD  = ("Consolas", 22, "bold")
FONT_TITLE = ("Consolas", 13, "bold")
FONT_BODY  = ("Consolas", 10)
FONT_SMALL = ("Consolas", 9)
FONT_MONO  = ("Courier New", 10)

# ══════════════════════════════════════════════════════════════════
#  DATABASE LAYER
# ══════════════════════════════════════════════════════════════════
class DB:
    def __init__(self):
        self.conn = sqlite3.connect(DB_FILE)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()
        self._seed_defaults()

    def _init_schema(self):
        c = self.conn.cursor()
        c.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT    UNIQUE NOT NULL,
            password TEXT    NOT NULL,
            role     TEXT    NOT NULL DEFAULT 'staff',
            created  TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS categories (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            sku         TEXT    UNIQUE NOT NULL,
            name        TEXT    NOT NULL,
            category_id INTEGER REFERENCES categories(id),
            quantity    INTEGER NOT NULL DEFAULT 0,
            threshold   INTEGER NOT NULL DEFAULT 5,
            unit_price  REAL    NOT NULL DEFAULT 0.0,
            location    TEXT,
            notes       TEXT,
            created     TEXT    NOT NULL,
            updated     TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS transactions (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id   INTEGER REFERENCES items(id),
            user_id   INTEGER REFERENCES users(id),
            action    TEXT    NOT NULL,
            qty_delta INTEGER NOT NULL,
            qty_after INTEGER NOT NULL,
            timestamp TEXT    NOT NULL,
            note      TEXT
        );
        """)
        self.conn.commit()

    def _seed_defaults(self):
        c = self.conn.cursor()
        # default admin
        c.execute("SELECT id FROM users WHERE username='admin'")
        if not c.fetchone():
            pw = hashlib.sha256("admin123".encode()).hexdigest()
            now = _now()
            c.execute("INSERT INTO users(username,password,role,created) VALUES(?,?,?,?)",
                      ("admin", pw, "admin", now))
        # default staff
        c.execute("SELECT id FROM users WHERE username='staff'")
        if not c.fetchone():
            pw = hashlib.sha256("staff123".encode()).hexdigest()
            now = _now()
            c.execute("INSERT INTO users(username,password,role,created) VALUES(?,?,?,?)",
                      ("staff", pw, "staff", now))
        # categories
        for cat in ["Electronics","Office Supplies","Furniture","Networking","Tools","Consumables"]:
            c.execute("INSERT OR IGNORE INTO categories(name) VALUES(?)", (cat,))
        # sample items
        c.execute("SELECT COUNT(*) FROM items")
        if c.fetchone()[0] == 0:
            now = _now()
            samples = [
                ("SKU-001","Dell Monitor 27\"",1,14,3,320.00,"Shelf A1",""),
                ("SKU-002","HP LaserJet Toner",2,50,10,45.00,"Shelf B3","Compatible model: M404"),
                ("SKU-003","Cisco Switch 24-Port",4,8,2,850.00,"Rack 2",""),
                ("SKU-004","Ergonomic Chair",3,6,2,220.00,"Warehouse",""),
                ("SKU-005","USB-C Hub 7-in-1",1,30,8,35.00,"Shelf A2",""),
                ("SKU-006","A4 Paper Ream 500s",2,120,20,4.50,"Shelf C1",""),
                ("SKU-007","RJ45 Patch Cable 2m",4,200,25,1.20,"Bin 7",""),
                ("SKU-008","Screwdriver Set",5,3,2,18.00,"Tool Cabinet",""),
            ]
            for s in samples:
                c.execute("""INSERT INTO items(sku,name,category_id,quantity,threshold,
                             unit_price,location,notes,created,updated)
                             VALUES(?,?,?,?,?,?,?,?,?,?)""",
                          (*s, now, now))
        self.conn.commit()

    # ── auth ───────────────────────────────────────────────────────
    def authenticate(self, username, password):
        pw = hashlib.sha256(password.encode()).hexdigest()
        c = self.conn.cursor()
        c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, pw))
        return c.fetchone()

    # ── users ──────────────────────────────────────────────────────
    def get_users(self):
        return self.conn.execute("SELECT * FROM users ORDER BY username").fetchall()

    def add_user(self, username, password, role):
        pw = hashlib.sha256(password.encode()).hexdigest()
        self.conn.execute("INSERT INTO users(username,password,role,created) VALUES(?,?,?,?)",
                          (username, pw, role, _now()))
        self.conn.commit()

    def delete_user(self, uid):
        self.conn.execute("DELETE FROM users WHERE id=?", (uid,))
        self.conn.commit()

    def change_password(self, uid, new_pw):
        pw = hashlib.sha256(new_pw.encode()).hexdigest()
        self.conn.execute("UPDATE users SET password=? WHERE id=?", (pw, uid))
        self.conn.commit()

    # ── categories ────────────────────────────────────────────────
    def get_categories(self):
        return self.conn.execute("SELECT * FROM categories ORDER BY name").fetchall()

    def add_category(self, name):
        self.conn.execute("INSERT OR IGNORE INTO categories(name) VALUES(?)", (name,))
        self.conn.commit()

    # ── items ──────────────────────────────────────────────────────
    def get_items(self, search="", category_id=None):
        q = """SELECT i.*, c.name AS cat_name
               FROM items i
               LEFT JOIN categories c ON i.category_id = c.id
               WHERE 1=1"""
        params = []
        if search:
            q += " AND (i.name LIKE ? OR i.sku LIKE ? OR i.location LIKE ?)"
            params += [f"%{search}%"]*3
        if category_id:
            q += " AND i.category_id=?"
            params.append(category_id)
        q += " ORDER BY i.name"
        return self.conn.execute(q, params).fetchall()

    def get_item(self, item_id):
        return self.conn.execute(
            "SELECT i.*, c.name AS cat_name FROM items i LEFT JOIN categories c ON i.category_id=c.id WHERE i.id=?",
            (item_id,)).fetchone()

    def add_item(self, sku, name, cat_id, qty, threshold, price, loc, notes):
        now = _now()
        self.conn.execute("""INSERT INTO items(sku,name,category_id,quantity,threshold,
                             unit_price,location,notes,created,updated)
                             VALUES(?,?,?,?,?,?,?,?,?,?)""",
                          (sku, name, cat_id, qty, threshold, price, loc, notes, now, now))
        self.conn.commit()

    def update_item(self, item_id, sku, name, cat_id, qty, threshold, price, loc, notes):
        self.conn.execute("""UPDATE items SET sku=?,name=?,category_id=?,quantity=?,
                             threshold=?,unit_price=?,location=?,notes=?,updated=?
                             WHERE id=?""",
                          (sku, name, cat_id, qty, threshold, price, loc, notes, _now(), item_id))
        self.conn.commit()

    def delete_item(self, item_id):
        self.conn.execute("DELETE FROM items WHERE id=?", (item_id,))
        self.conn.commit()

    def adjust_stock(self, item_id, user_id, action, delta, note=""):
        item = self.get_item(item_id)
        new_qty = item["quantity"] + delta
        if new_qty < 0:
            raise ValueError("Stock cannot go below zero.")
        self.conn.execute("UPDATE items SET quantity=?, updated=? WHERE id=?",
                          (new_qty, _now(), item_id))
        self.conn.execute("""INSERT INTO transactions(item_id,user_id,action,qty_delta,qty_after,timestamp,note)
                             VALUES(?,?,?,?,?,?,?)""",
                          (item_id, user_id, action, delta, new_qty, _now(), note))
        self.conn.commit()
        return new_qty

    # ── transactions ──────────────────────────────────────────────
    def get_transactions(self, limit=200):
        return self.conn.execute("""
            SELECT t.*, i.name AS item_name, i.sku, u.username
            FROM transactions t
            LEFT JOIN items i ON t.item_id=i.id
            LEFT JOIN users u ON t.user_id=u.id
            ORDER BY t.timestamp DESC LIMIT ?""", (limit,)).fetchall()

    def get_low_stock(self):
        return self.conn.execute("""SELECT i.*, c.name AS cat_name
            FROM items i LEFT JOIN categories c ON i.category_id=c.id
            WHERE i.quantity <= i.threshold ORDER BY i.quantity""").fetchall()

    def close(self):
        self.conn.close()


def _now():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ══════════════════════════════════════════════════════════════════
#  EXPORT HELPERS
# ══════════════════════════════════════════════════════════════════
def export_excel(items, filename="inventory_report.xlsx"):
    if not EXCEL_OK:
        return False, "openpyxl not installed. Run: pip install openpyxl"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    hdr_fill = PatternFill("solid", fgColor="238636")
    hdr_font = Font(bold=True, color="FFFFFF", name="Consolas")
    body_font = Font(name="Consolas", size=10)
    alt_fill  = PatternFill("solid", fgColor="161B22")
    thin = Border(
        left=Side(style="thin",color="30363D"),
        right=Side(style="thin",color="30363D"),
        top=Side(style="thin",color="30363D"),
        bottom=Side(style="thin",color="30363D"),
    )

    headers = ["SKU","Name","Category","Qty","Threshold","Unit Price","Location","Notes","Last Updated"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for i, row in enumerate(items, 2):
        vals = [row["sku"], row["name"], row["cat_name"] or "",
                row["quantity"], row["threshold"], row["unit_price"],
                row["location"] or "", row["notes"] or "", row["updated"]]
        ws.append(vals)
        for cell in ws[i]:
            cell.font = body_font
            cell.border = thin
            if i % 2 == 0:
                cell.fill = alt_fill

        # colour low-stock cells
        qty_cell = ws.cell(i, 4)
        thr_cell = ws.cell(i, 5)
        if row["quantity"] <= row["threshold"]:
            qty_cell.fill = PatternFill("solid", fgColor="3D1F1F")
            qty_cell.font = Font(name="Consolas", size=10, color="DA3633", bold=True)

    col_widths = [12,30,16,8,12,12,16,24,22]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w

    # summary sheet
    ws2 = wb.create_sheet("Summary")
    total_val = sum(r["quantity"]*r["unit_price"] for r in items)
    low = sum(1 for r in items if r["quantity"] <= r["threshold"])
    ws2.append(["StockSentinel Pro – Inventory Summary"])
    ws2.append([f"Generated: {_now()}"])
    ws2.append([])
    ws2.append(["Total SKUs", len(items)])
    ws2.append(["Total Units", sum(r["quantity"] for r in items)])
    ws2.append(["Total Value (USD)", round(total_val,2)])
    ws2.append(["Low-Stock Items", low])

    wb.save(filename)
    return True, filename


def export_pdf(items, filename="inventory_report.pdf"):
    if not PDF_OK:
        return False, "reportlab not installed. Run: pip install reportlab"
    doc = SimpleDocTemplate(filename, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle("title", fontSize=18, leading=24,
                                 textColor=colors.HexColor("#238636"),
                                 fontName="Courier-Bold")
    sub_style   = ParagraphStyle("sub", fontSize=9, leading=12,
                                 textColor=colors.HexColor("#8B949E"),
                                 fontName="Courier")

    story.append(Paragraph("StockSentinel Pro", title_style))
    story.append(Paragraph(f"Inventory Report — {_now()}", sub_style))
    story.append(Spacer(1, 0.4*cm))

    headers = ["SKU","Name","Category","Qty","Threshold","Price","Location"]
    data    = [headers]
    for r in items:
        data.append([r["sku"], r["name"][:28], r["cat_name"] or "",
                     str(r["quantity"]), str(r["threshold"]),
                     f"${r['unit_price']:.2f}", r["location"] or ""])

    t = Table(data, repeatRows=1, hAlign="LEFT")
    ts = TableStyle([
        ("BACKGROUND",  (0,0),(-1,0), colors.HexColor("#238636")),
        ("TEXTCOLOR",   (0,0),(-1,0), colors.white),
        ("FONTNAME",    (0,0),(-1,0), "Courier-Bold"),
        ("FONTSIZE",    (0,0),(-1,-1), 8),
        ("FONTNAME",    (0,1),(-1,-1), "Courier"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
         [colors.HexColor("#161B22"), colors.HexColor("#1C2128")]),
        ("TEXTCOLOR",   (0,1),(-1,-1), colors.HexColor("#E6EDF3")),
        ("GRID",        (0,0),(-1,-1), 0.3, colors.HexColor("#30363D")),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ])
    # highlight low stock
    for i, r in enumerate(items, 1):
        if r["quantity"] <= r["threshold"]:
            ts.add("TEXTCOLOR",  (3,i),(3,i), colors.HexColor("#DA3633"))
            ts.add("FONTNAME",   (3,i),(3,i), "Courier-Bold")
    t.setStyle(ts)
    story.append(t)

    story.append(Spacer(1, 0.4*cm))
    total_val = sum(r["quantity"]*r["unit_price"] for r in items)
    low_count = sum(1 for r in items if r["quantity"] <= r["threshold"])
    summary = [
        ["Total SKUs", str(len(items))],
        ["Total Units", str(sum(r["quantity"] for r in items))],
        ["Total Value", f"${total_val:,.2f}"],
        ["Low-Stock Items", str(low_count)],
    ]
    st = Table(summary, colWidths=[5*cm, 4*cm])
    st.setStyle(TableStyle([
        ("BACKGROUND",  (0,0),(-1,-1), colors.HexColor("#21262D")),
        ("TEXTCOLOR",   (0,0),(-1,-1), colors.HexColor("#E6EDF3")),
        ("FONTNAME",    (0,0),(0,-1),  "Courier-Bold"),
        ("FONTNAME",    (1,0),(1,-1),  "Courier"),
        ("FONTSIZE",    (0,0),(-1,-1), 9),
        ("GRID",        (0,0),(-1,-1), 0.3, colors.HexColor("#30363D")),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    story.append(st)
    doc.build(story)
    return True, filename


# ══════════════════════════════════════════════════════════════════
#  REUSABLE UI WIDGETS
# ══════════════════════════════════════════════════════════════════
def styled_btn(parent, text, cmd, color=None, width=18, small=False):
    bg  = color or C["accent"]
    hbg = C["accent_h"] if bg == C["accent"] else bg
    fnt = FONT_SMALL if small else FONT_BODY
    btn = tk.Button(parent, text=text, command=cmd, bg=bg, fg=C["white"],
                    font=fnt, relief="flat", cursor="hand2",
                    activebackground=hbg, activeforeground=C["white"],
                    width=width, pady=4)
    return btn

def label(parent, text, font=None, fg=None, **kw):
    return tk.Label(parent, text=text,
                    font=font or FONT_BODY,
                    fg=fg or C["text"],
                    bg=C["surface"],
                    **kw)

def sep(parent):
    return tk.Frame(parent, height=1, bg=C["border"])

def card(parent, **kw):
    f = tk.Frame(parent, bg=C["surface"], relief="flat", bd=0, **kw)
    return f

def entry_field(parent, var, width=28, show=None):
    e = tk.Entry(parent, textvariable=var, width=width,
                 bg=C["surface2"], fg=C["text"], insertbackground=C["text"],
                 relief="flat", font=FONT_BODY, show=show or "")
    e.config(highlightthickness=1, highlightbackground=C["border"],
             highlightcolor=C["highlight"])
    return e


# ══════════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ══════════════════════════════════════════════════════════════════
class LoginWindow:
    def __init__(self, root, db, on_login):
        self.root = root
        self.db   = db
        self.on_login = on_login
        root.title("StockSentinel Pro — Login")
        root.configure(bg=C["bg"])
        root.resizable(False, False)
        self._build()
        root.update_idletasks()
        w,h = 420, 480
        x = (root.winfo_screenwidth()-w)//2
        y = (root.winfo_screenheight()-h)//2
        root.geometry(f"{w}x{h}+{x}+{y}")

    def _build(self):
        r = self.root
        # top bar
        top = tk.Frame(r, bg=C["accent"], height=4)
        top.pack(fill="x")

        body = tk.Frame(r, bg=C["bg"])
        body.pack(fill="both", expand=True, padx=40, pady=30)

        tk.Label(body, text="◈", font=("Consolas",42), fg=C["accent"],
                 bg=C["bg"]).pack(pady=(10,0))
        tk.Label(body, text="StockSentinel Pro", font=("Consolas",20,"bold"),
                 fg=C["text"], bg=C["bg"]).pack()
        tk.Label(body, text="Inventory Management System", font=FONT_SMALL,
                 fg=C["muted"], bg=C["bg"]).pack(pady=(0,24))

        frm = tk.Frame(body, bg=C["surface"], relief="flat",
                       highlightthickness=1,
                       highlightbackground=C["border"])
        frm.pack(fill="x", pady=4, ipady=20, ipadx=20)

        for txt, attr, show in [("Username","uvar",None),("Password","pvar","●")]:
            tk.Label(frm, text=txt, font=FONT_SMALL, fg=C["muted"],
                     bg=C["surface"], anchor="w").pack(fill="x", padx=20, pady=(12,2))
            var = tk.StringVar()
            setattr(self, attr, var)
            e = entry_field(frm, var, width=30, show=show)
            e.pack(fill="x", padx=20, ipady=6)
            if attr == "pvar":
                e.bind("<Return>", lambda _: self._login())

        tk.Label(frm, text="", bg=C["surface"]).pack(pady=4)
        styled_btn(frm, "  LOGIN", self._login, width=30).pack(padx=20, pady=(0,12))

        tk.Label(body, text="Default: admin/admin123  |  staff/staff123",
                 font=FONT_SMALL, fg=C["muted"], bg=C["bg"]).pack(pady=8)

    def _login(self):
        user = self.db.authenticate(self.uvar.get().strip(), self.pvar.get())
        if user:
            self.on_login(dict(user))
        else:
            messagebox.showerror("Auth Failed", "Invalid username or password.", parent=self.root)
            self.pvar.set("")


# ══════════════════════════════════════════════════════════════════
#  MAIN APPLICATION WINDOW
# ══════════════════════════════════════════════════════════════════
class App:
    def __init__(self, root, db, user):
        self.root = root
        self.db   = db
        self.user = user          # dict: id, username, role
        self.is_admin = user["role"] == "admin"
        self._search_after = None

        root.title(f"StockSentinel Pro  ·  {user['username']} [{user['role'].upper()}]")
        root.configure(bg=C["bg"])
        root.state("zoomed")
        root.protocol("WM_DELETE_WINDOW", self._quit)

        self._build_ui()
        self.show_inventory()

    # ── SKELETON ──────────────────────────────────────────────────
    def _build_ui(self):
        # top bar
        topbar = tk.Frame(self.root, bg=C["surface"], height=52)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        tk.Label(topbar, text="◈  StockSentinel Pro",
                 font=("Consolas",15,"bold"), fg=C["accent"],
                 bg=C["surface"]).pack(side="left", padx=18)

        self.clock_lbl = tk.Label(topbar, text="", font=FONT_SMALL,
                                  fg=C["muted"], bg=C["surface"])
        self.clock_lbl.pack(side="right", padx=16)
        self._tick()

        tk.Label(topbar, text=f"  {self.user['username']}  [{self.user['role'].upper()}]",
                 font=FONT_SMALL, fg=C["muted"], bg=C["surface"]).pack(side="right")

        # separator
        sep(self.root).pack(fill="x")

        # main layout
        layout = tk.Frame(self.root, bg=C["bg"])
        layout.pack(fill="both", expand=True)

        # sidebar
        self.sidebar = tk.Frame(layout, bg=C["surface"], width=200)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        sep_v = tk.Frame(layout, width=1, bg=C["border"])
        sep_v.pack(side="left", fill="y")

        # content
        self.content = tk.Frame(layout, bg=C["bg"])
        self.content.pack(side="left", fill="both", expand=True)

    def _build_sidebar(self):
        sb = self.sidebar
        tk.Label(sb, text="NAVIGATION", font=("Consolas",8,"bold"),
                 fg=C["muted"], bg=C["surface"], anchor="w").pack(fill="x", padx=14, pady=(18,6))

        nav_items = [
            ("▤  Inventory",    self.show_inventory),
            ("⊕  Transactions", self.show_transactions),
            ("⚠  Low Stock",    self.show_low_stock),
        ]
        if self.is_admin:
            nav_items += [
                ("◉  Users",       self.show_users),
                ("⬡  Categories",  self.show_categories),
            ]
        nav_items.append(("⤻  Logout",     self._logout))

        self.nav_btns = []
        for label_text, cmd in nav_items:
            btn = tk.Button(sb, text=label_text, command=cmd,
                            bg=C["surface"], fg=C["text"],
                            font=FONT_BODY, relief="flat", anchor="w",
                            cursor="hand2", padx=14, pady=10,
                            activebackground=C["surface2"],
                            activeforeground=C["accent"])
            btn.pack(fill="x")
            self.nav_btns.append(btn)

        sep(sb).pack(fill="x", pady=8)
        tk.Label(sb, text="EXPORT", font=("Consolas",8,"bold"),
                 fg=C["muted"], bg=C["surface"], anchor="w").pack(fill="x", padx=14, pady=(0,6))

        styled_btn(sb, "⤓ Export Excel", self._export_excel, width=22, small=True).pack(padx=10, pady=3)
        styled_btn(sb, "⤓ Export PDF",   self._export_pdf,   width=22, small=True,
                   color=C["info"]).pack(padx=10, pady=3)

    def _set_active_nav(self, idx):
        for i, btn in enumerate(self.nav_btns):
            if i == idx:
                btn.config(bg=C["surface2"], fg=C["accent"])
            else:
                btn.config(bg=C["surface"], fg=C["text"])

    def _clear_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    def _tick(self):
        self.clock_lbl.config(text=datetime.datetime.now().strftime("%a %d %b %Y  %H:%M:%S"))
        self.root.after(1000, self._tick)

    # ── INVENTORY VIEW ────────────────────────────────────────────
    def show_inventory(self):
        self._set_active_nav(0)
        self._clear_content()
        f = self.content

        # toolbar
        tb = tk.Frame(f, bg=C["bg"])
        tb.pack(fill="x", padx=16, pady=12)

        tk.Label(tb, text="Inventory", font=FONT_HEAD,
                 fg=C["text"], bg=C["bg"]).pack(side="left")

        if self.is_admin:
            styled_btn(tb, "+ Add Item", self._dlg_add_item, width=14).pack(side="right", padx=4)

        # search row
        sr = tk.Frame(f, bg=C["bg"])
        sr.pack(fill="x", padx=16, pady=(0,8))

        self.search_var = tk.StringVar()
        se = entry_field(sr, self.search_var, width=36)
        se.pack(side="left", ipady=5)
        tk.Label(sr, text=" 🔍 Search SKU / Name / Location",
                 font=FONT_SMALL, fg=C["muted"], bg=C["bg"]).pack(side="left")

        # category filter
        tk.Label(sr, text="  Filter:", font=FONT_SMALL, fg=C["muted"], bg=C["bg"]).pack(side="left", padx=(20,4))
        self.cat_filter = tk.StringVar(value="All")
        cats = ["All"] + [c["name"] for c in self.db.get_categories()]
        om = ttk.Combobox(sr, textvariable=self.cat_filter, values=cats,
                          width=16, state="readonly", font=FONT_SMALL)
        om.pack(side="left")
        om.bind("<<ComboboxSelected>>", lambda _: self._refresh_inventory())

        self.search_var.trace_add("write", self._search_delayed)

        # treeview
        cols = ("SKU","Name","Category","Qty","Threshold","Price","Location","Updated")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", selectmode="browse")
        widths    = [90, 240, 120, 60, 80, 80, 120, 150]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col, command=lambda c=col: self._sort_tree(c))
            self.tree.column(col, width=w, anchor="center" if col in ("Qty","Threshold","Price") else "w")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=C["surface"],
                        foreground=C["text"],
                        rowheight=28,
                        fieldbackground=C["surface"],
                        font=FONT_BODY)
        style.configure("Treeview.Heading",
                        background=C["surface2"],
                        foreground=C["muted"],
                        font=FONT_SMALL,
                        relief="flat")
        style.map("Treeview", background=[("selected", C["surface2"])],
                              foreground=[("selected", C["accent"])])

        self.tree.tag_configure("low",    background="#2D1B1B", foreground=C["danger"])
        self.tree.tag_configure("ok",     background=C["surface"])
        self.tree.tag_configure("alt",    background=C["tag_bg"])

        sb = ttk.Scrollbar(f, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(16,0), pady=(0,12))
        sb.pack(side="left", fill="y", pady=(0,12))

        # action bar
        ab = tk.Frame(f, bg=C["bg"])
        ab.pack(fill="x", padx=16, pady=(0,12))
        styled_btn(ab, "▲ Check In",  lambda: self._stock_action("in"),  width=14).pack(side="left", padx=4)
        styled_btn(ab, "▼ Check Out", lambda: self._stock_action("out"),
                   color=C["warning"], width=14).pack(side="left", padx=4)
        if self.is_admin:
            styled_btn(ab, "✎ Edit",   self._dlg_edit_item,  color=C["info"],   width=12).pack(side="left", padx=4)
            styled_btn(ab, "✕ Delete", self._delete_item,    color=C["danger"], width=12).pack(side="left", padx=4)

        self.status_var = tk.StringVar()
        tk.Label(ab, textvariable=self.status_var, font=FONT_SMALL,
                 fg=C["muted"], bg=C["bg"]).pack(side="right", padx=8)

        self._refresh_inventory()

    def _search_delayed(self, *_):
        if self._search_after:
            self.root.after_cancel(self._search_after)
        self._search_after = self.root.after(250, self._refresh_inventory)

    def _refresh_inventory(self):
        self.tree.delete(*self.tree.get_children())
        search = self.search_var.get().strip()
        cat_name = self.cat_filter.get()
        cat_id = None
        if cat_name != "All":
            for c in self.db.get_categories():
                if c["name"] == cat_name:
                    cat_id = c["id"]
        items = self.db.get_items(search, cat_id)
        for i, row in enumerate(items):
            tag = "low" if row["quantity"] <= row["threshold"] else ("alt" if i%2 else "ok")
            self.tree.insert("", "end", iid=str(row["id"]), tags=(tag,),
                             values=(row["sku"], row["name"], row["cat_name"] or "",
                                     row["quantity"], row["threshold"],
                                     f"${row['unit_price']:.2f}",
                                     row["location"] or "", row["updated"]))
        low = sum(1 for r in items if r["quantity"] <= r["threshold"])
        self.status_var.set(f"{len(items)} items  |  ⚠ {low} low-stock")

    def _sort_tree(self, col):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        try:
            items.sort(key=lambda t: float(t[0].replace("$","")) if t[0].replace("$","").replace(".","").isdigit() else t[0])
        except:
            items.sort()
        for idx, (_, k) in enumerate(items):
            self.tree.move(k, "", idx)

    def _selected_item_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select an item first.")
            return None
        return int(sel[0])

    def _stock_action(self, direction):
        iid = self._selected_item_id()
        if not iid: return
        item = self.db.get_item(iid)
        sign = 1 if direction == "in" else -1
        label_txt = "Check In (add qty)" if direction == "in" else "Check Out (remove qty)"
        qty = simpledialog.askinteger(label_txt, f"Quantity to {'add to' if direction=='in' else 'remove from'} '{item['name']}':",
                                     minvalue=1, maxvalue=9999, parent=self.root)
        if qty is None: return
        note = simpledialog.askstring("Note (optional)", "Add a note for this transaction:", parent=self.root) or ""
        try:
            new_qty = self.db.adjust_stock(iid, self.user["id"], direction, sign*qty, note)
            messagebox.showinfo("Done", f"Stock updated. New quantity: {new_qty}", parent=self.root)
        except ValueError as e:
            messagebox.showerror("Error", str(e), parent=self.root)
        self._refresh_inventory()

    # ── ADD / EDIT ITEM DIALOG ────────────────────────────────────
    def _item_dialog(self, title, item=None):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.configure(bg=C["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.update_idletasks()
        w,h = 480,520
        x = (dlg.winfo_screenwidth()-w)//2
        y = (dlg.winfo_screenheight()-h)//2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

        fields = {}
        cats   = self.db.get_categories()
        cat_names = [c["name"] for c in cats]

        rows = [
            ("SKU",        "sku",        ""),
            ("Name",       "name",       ""),
            ("Quantity",   "quantity",   "0"),
            ("Threshold",  "threshold",  "5"),
            ("Unit Price", "unit_price", "0.00"),
            ("Location",   "location",   ""),
            ("Notes",      "notes",      ""),
        ]
        tk.Label(dlg, text=title, font=FONT_TITLE, fg=C["accent"],
                 bg=C["surface"]).pack(pady=16)

        for lbl, key, default in rows:
            row = tk.Frame(dlg, bg=C["surface"])
            row.pack(fill="x", padx=30, pady=4)
            tk.Label(row, text=lbl, font=FONT_SMALL, fg=C["muted"],
                     bg=C["surface"], width=12, anchor="w").pack(side="left")
            var = tk.StringVar(value=(str(item[key]) if item and item[key] is not None else default))
            e = entry_field(row, var, width=30)
            e.pack(side="left", ipady=4)
            fields[key] = var

        # category dropdown
        cat_row = tk.Frame(dlg, bg=C["surface"])
        cat_row.pack(fill="x", padx=30, pady=4)
        tk.Label(cat_row, text="Category", font=FONT_SMALL, fg=C["muted"],
                 bg=C["surface"], width=12, anchor="w").pack(side="left")
        cat_var = tk.StringVar()
        if item:
            cat_var.set(item["cat_name"] or cat_names[0])
        else:
            cat_var.set(cat_names[0])
        om = ttk.Combobox(cat_row, textvariable=cat_var, values=cat_names,
                          width=28, state="readonly", font=FONT_SMALL)
        om.pack(side="left")

        result = {"submitted": False}
        def submit():
            try:
                sku   = fields["sku"].get().strip()
                name  = fields["name"].get().strip()
                qty   = int(fields["quantity"].get())
                thr   = int(fields["threshold"].get())
                price = float(fields["unit_price"].get())
                loc   = fields["location"].get().strip()
                notes = fields["notes"].get().strip()
                cat   = next(c["id"] for c in cats if c["name"] == cat_var.get())
                if not sku or not name:
                    raise ValueError("SKU and Name are required.")
                result.update(dict(sku=sku, name=name, cat_id=cat,
                                   qty=qty, threshold=thr, price=price,
                                   loc=loc, notes=notes, submitted=True))
                dlg.destroy()
            except (ValueError, StopIteration) as e:
                messagebox.showerror("Validation", str(e), parent=dlg)

        btn_row = tk.Frame(dlg, bg=C["surface"])
        btn_row.pack(pady=16)
        styled_btn(btn_row, "Save", submit, width=14).pack(side="left", padx=8)
        styled_btn(btn_row, "Cancel", dlg.destroy, color=C["danger"], width=10).pack(side="left")

        dlg.wait_window()
        return result

    def _dlg_add_item(self):
        r = self._item_dialog("Add New Item")
        if r["submitted"]:
            try:
                self.db.add_item(r["sku"], r["name"], r["cat_id"], r["qty"],
                                 r["threshold"], r["price"], r["loc"], r["notes"])
                self._refresh_inventory()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "SKU already exists.")

    def _dlg_edit_item(self):
        iid = self._selected_item_id()
        if not iid: return
        item = dict(self.db.get_item(iid))
        r = self._item_dialog("Edit Item", item)
        if r["submitted"]:
            self.db.update_item(iid, r["sku"], r["name"], r["cat_id"], r["qty"],
                                r["threshold"], r["price"], r["loc"], r["notes"])
            self._refresh_inventory()

    def _delete_item(self):
        iid = self._selected_item_id()
        if not iid: return
        item = self.db.get_item(iid)
        if messagebox.askyesno("Confirm Delete",
                               f"Delete '{item['name']}'?\nThis cannot be undone.",
                               parent=self.root):
            self.db.delete_item(iid)
            self._refresh_inventory()

    # ── TRANSACTIONS VIEW ─────────────────────────────────────────
    def show_transactions(self):
        self._set_active_nav(1)
        self._clear_content()
        f = self.content

        tk.Label(f, text="Transaction Log", font=FONT_HEAD,
                 fg=C["text"], bg=C["bg"]).pack(anchor="w", padx=16, pady=12)

        cols = ("Time","SKU","Item","User","Action","Δ Qty","After","Note")
        tree = ttk.Treeview(f, columns=cols, show="headings")
        widths = [150, 80, 220, 100, 80, 60, 60, 200]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center" if col in ("Δ Qty","After") else "w")

        tree.tag_configure("in",  foreground=C["accent"])
        tree.tag_configure("out", foreground=C["warning"])

        rows = self.db.get_transactions()
        for row in rows:
            delta = f"+{row['qty_delta']}" if row["qty_delta"] > 0 else str(row["qty_delta"])
            tree.insert("", "end", tags=(row["action"],),
                        values=(row["timestamp"], row["sku"] or "",
                                row["item_name"] or "(deleted)",
                                row["username"] or "",
                                row["action"].upper(), delta,
                                row["qty_after"], row["note"] or ""))

        sb = ttk.Scrollbar(f, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True, padx=(16,0), pady=(0,16))
        sb.pack(side="left", fill="y", pady=(0,16))

    # ── LOW STOCK VIEW ────────────────────────────────────────────
    def show_low_stock(self):
        self._set_active_nav(2)
        self._clear_content()
        f = self.content

        tk.Label(f, text="⚠  Low Stock Alerts", font=FONT_HEAD,
                 fg=C["danger"], bg=C["bg"]).pack(anchor="w", padx=16, pady=12)

        items = self.db.get_low_stock()
        if not items:
            tk.Label(f, text="✔ All items are above threshold.",
                     font=FONT_TITLE, fg=C["accent"], bg=C["bg"]).pack(pady=40)
            return

        cols = ("SKU","Name","Category","Qty","Threshold","Location")
        tree = ttk.Treeview(f, columns=cols, show="headings")
        widths = [90, 260, 120, 60, 80, 140]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center" if col in ("Qty","Threshold") else "w")
        tree.tag_configure("critical", background="#3D1212", foreground=C["danger"])
        for row in items:
            tree.insert("", "end", tags=("critical",),
                        values=(row["sku"], row["name"], row["cat_name"] or "",
                                row["quantity"], row["threshold"], row["location"] or ""))
        tree.pack(fill="both", expand=True, padx=16, pady=(0,16))

    # ── USERS VIEW (admin) ────────────────────────────────────────
    def show_users(self):
        if not self.is_admin: return
        self._set_active_nav(3)
        self._clear_content()
        f = self.content

        tb = tk.Frame(f, bg=C["bg"])
        tb.pack(fill="x", padx=16, pady=12)
        tk.Label(tb, text="User Management", font=FONT_HEAD,
                 fg=C["text"], bg=C["bg"]).pack(side="left")
        styled_btn(tb, "+ Add User", self._dlg_add_user, width=14).pack(side="right")

        cols = ("ID","Username","Role","Created")
        tree = ttk.Treeview(f, columns=cols, show="headings", selectmode="browse")
        for col, w in zip(cols, [50, 200, 100, 180]):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="w")

        def refresh():
            tree.delete(*tree.get_children())
            for u in self.db.get_users():
                tree.insert("", "end", iid=str(u["id"]),
                            values=(u["id"], u["username"], u["role"], u["created"]))

        refresh()
        tree.pack(fill="both", expand=True, padx=16, pady=(0,8))

        ab = tk.Frame(f, bg=C["bg"])
        ab.pack(fill="x", padx=16, pady=(0,12))
        def del_user():
            sel = tree.selection()
            if not sel: return
            uid = int(sel[0])
            if uid == self.user["id"]:
                messagebox.showwarning("Cannot Delete", "You cannot delete your own account.")
                return
            if messagebox.askyesno("Confirm", "Delete selected user?"):
                self.db.delete_user(uid)
                refresh()
        def chg_pw():
            sel = tree.selection()
            if not sel: return
            uid = int(sel[0])
            pw = simpledialog.askstring("Change Password", "New password:", show="*", parent=self.root)
            if pw:
                self.db.change_password(uid, pw)
                messagebox.showinfo("Done", "Password changed.")

        styled_btn(ab, "✕ Delete User",    del_user, color=C["danger"], width=16).pack(side="left", padx=4)
        styled_btn(ab, "🔑 Change Password", chg_pw, color=C["info"],  width=18).pack(side="left", padx=4)

    def _dlg_add_user(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Add User")
        dlg.configure(bg=C["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("380x300+500+300")

        uvar  = tk.StringVar()
        pvar  = tk.StringVar()
        rvar  = tk.StringVar(value="staff")

        tk.Label(dlg, text="Add New User", font=FONT_TITLE, fg=C["accent"], bg=C["surface"]).pack(pady=14)
        for lbl, var, show in [("Username", uvar, None), ("Password", pvar, "●")]:
            r = tk.Frame(dlg, bg=C["surface"])
            r.pack(fill="x", padx=24, pady=4)
            tk.Label(r, text=lbl, font=FONT_SMALL, fg=C["muted"], bg=C["surface"], width=10, anchor="w").pack(side="left")
            entry_field(r, var, show=show).pack(side="left", ipady=4)

        r2 = tk.Frame(dlg, bg=C["surface"])
        r2.pack(fill="x", padx=24, pady=4)
        tk.Label(r2, text="Role", font=FONT_SMALL, fg=C["muted"], bg=C["surface"], width=10, anchor="w").pack(side="left")
        ttk.Combobox(r2, textvariable=rvar, values=["staff","admin"],
                     state="readonly", width=16, font=FONT_SMALL).pack(side="left")

        def save():
            u,p,ro = uvar.get().strip(), pvar.get().strip(), rvar.get()
            if not u or not p:
                messagebox.showerror("Validation", "Username and password required.", parent=dlg)
                return
            try:
                self.db.add_user(u, p, ro)
                dlg.destroy()
                self.show_users()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "Username already exists.", parent=dlg)

        btn_r = tk.Frame(dlg, bg=C["surface"])
        btn_r.pack(pady=14)
        styled_btn(btn_r, "Create", save, width=12).pack(side="left", padx=6)
        styled_btn(btn_r, "Cancel", dlg.destroy, color=C["danger"], width=10).pack(side="left")
        dlg.wait_window()

    # ── CATEGORIES VIEW (admin) ───────────────────────────────────
    def show_categories(self):
        if not self.is_admin: return
        self._set_active_nav(4)
        self._clear_content()
        f = self.content

        tb = tk.Frame(f, bg=C["bg"])
        tb.pack(fill="x", padx=16, pady=12)
        tk.Label(tb, text="Categories", font=FONT_HEAD, fg=C["text"], bg=C["bg"]).pack(side="left")

        lbox = tk.Listbox(f, bg=C["surface"], fg=C["text"], font=FONT_BODY,
                          selectbackground=C["surface2"], selectforeground=C["accent"],
                          relief="flat", highlightthickness=1,
                          highlightbackground=C["border"], width=40, height=20)

        def refresh():
            lbox.delete(0, "end")
            for c in self.db.get_categories():
                lbox.insert("end", f"  {c['name']}")

        refresh()
        lbox.pack(padx=16, pady=(0,8))

        ab = tk.Frame(f, bg=C["bg"])
        ab.pack(padx=16, anchor="w")
        def add_cat():
            name = simpledialog.askstring("Add Category", "Category name:", parent=self.root)
            if name:
                self.db.add_category(name.strip())
                refresh()
        styled_btn(ab, "+ Add Category", add_cat, width=16).pack(side="left")

    # ── EXPORT ────────────────────────────────────────────────────
    def _export_excel(self):
        items = self.db.get_items()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"inventory_{ts}.xlsx"
        ok, msg = export_excel(items, fname)
        if ok:
            messagebox.showinfo("Exported", f"Excel report saved:\n{os.path.abspath(fname)}")
        else:
            messagebox.showerror("Export Failed", msg)

    def _export_pdf(self):
        items = self.db.get_items()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"inventory_{ts}.pdf"
        ok, msg = export_pdf(items, fname)
        if ok:
            messagebox.showinfo("Exported", f"PDF report saved:\n{os.path.abspath(fname)}")
        else:
            messagebox.showerror("Export Failed", msg)

    # ── QUIT / LOGOUT ─────────────────────────────────────────────
    def _logout(self):
        if messagebox.askyesno("Logout", "Return to login screen?"):
            self._clear_content()
            for w in self.root.winfo_children():
                w.destroy()
            self.db  # keep db open
            start_login(self.root, self.db)

    def _quit(self):
        self.db.close()
        self.root.destroy()


# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════
def start_login(root, db):
    root.state("normal")
    root.resizable(False, False)

    def on_login(user):
        for w in root.winfo_children():
            w.destroy()
        App(root, db, user)

    LoginWindow(root, db, on_login)


def main():
    print("""
  ╔═══════════════════════════════════════╗
  ║       StockSentinel Pro v1.0          ║
  ║   Comprehensive Inventory Manager     ║
  ╠═══════════════════════════════════════╣
  ║  DB file   : stocksentinel.db         ║
  ║  Default logins:                      ║
  ║    admin  / admin123                  ║
  ║    staff  / staff123                  ║
  ╠═══════════════════════════════════════╣
  ║  Optional libs for export:            ║
  ║    pip install openpyxl               ║
  ║    pip install reportlab              ║
  ╚═══════════════════════════════════════╝
""")
    db   = DB()
    root = tk.Tk()
    root.iconname("StockSentinel")
    start_login(root, db)
    root.mainloop()


if __name__ == "__main__":
    main()